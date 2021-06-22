import json
import yaml
import datetime
from pprint import pprint

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from nornir import InitNornir
from nornir_utils.plugins.functions import print_result
from nornir.core.task import Task, Result
from nornir_netmiko import netmiko_send_command
from nornir_salt import ResultSerializer

from parsing import CiscoParse


def get_hosts_file(excel_file):
    try:
        book = load_workbook(excel_file)
        sheet = book['hosts']
    except Exception as e:
        print('Excel File Error occoured...\n %s' % (e,))
        exit()

    hosts = dict()
    for row in sheet.rows:
        key = row[0].value
        if key:
            hosts[key] = {
                'hostname': row[1].value.strip(),
                'username': row[2].value.strip(),
                'password': row[3].value.strip(),
                'platform': row[4].value.strip(),
                'port': row[5].value,
                'data': {
                    'vendors': row[6].value.strip(),
                    'role': row[7].value.strip(),
                    'site': row[8].value.strip(),
                },
            }

    book.close()
    # remove first row of excel sheet.
    if hosts['id']:
        del hosts['id']

    inventory_hosts = {'hosts': hosts}
    return inventory_hosts


# use only SimpleInventory plugins..
def hosts_to_yaml(hosts):
    json_data = json.dumps(hosts)
    yaml_data = yaml.load(json_data, Loader=yaml.SafeLoader)
    f = open('hosts.yaml', 'w')
    f.write(yaml.dump(yaml_data, default_flow_style=False))
    f.close()


def device_ip(task: Task) -> Result:
    return Result(
        host=task.host,
        result=f'{task.host.hostname}'
    )


# no more use. it was replacee netmiko_send_commands.
def ios_group_task(task: Task) -> Result:
    task.run(
        task=netmiko_send_command,
        command_string='show hardware',
        name='HW_info',
    )
    task.run(
        task=netmiko_send_command,
        command_string='show env all',
        name='HW_check1',
    )
    task.run(
        task=netmiko_send_command,
        command_string='show env',
        name='HW_check2',
    )
    task.run(
        task=netmiko_send_command,
        command_string='show processes cpu',
        name='CPU_info',
    )
    task.run(
        task=netmiko_send_command,
        command_string='show processes mem',
        name='MEM_info',
    )
    task.run(
        task=netmiko_send_command,
        command_string='show run | inc hostname',
        name='Hostname',
    )

    task.run(task=device_ip, name='device_ip')
    return Result(host=task.host)


def netmiko_send_commands(task, commands, **kwargs):
    for command in commands:
        task.run(
            task=netmiko_send_command,
            command_string=command,
            name=command,
            **kwargs.get('netmiko_kwargs', {})
        )
    task.run(task=device_ip)
    return Result(host=task.host)


def ios_pm(inventory_hosts, **kwargs):
    nr = InitNornir(
        runner={
            'plugin': 'threaded',
            'options': {
                'num_workers': 30,
            },
        },
        inventory={
            # 'plugin': 'SimpleInventory',
            'plugin': 'DictInventory',
            'options': {
                # 'host_file': 'hosts.yaml',
                'hosts': inventory_hosts['hosts'],
                'groups': {},
                'defaults': {},
            }
        }
    )
    vendors = kwargs['vendors']
    commands = kwargs['commands']
    site = ''
    role = ''
    if kwargs['site']:
        site = kwargs['site']
    if kwargs['role']:
        role = kwargs['role']
    if not commands:
        print('Need to commands...')

    if site and role:
        nr = nr.filter(vendors='cisco', site=site, role=role)
    else:
        if site:
            nr = nr.filter(vendors='cisco', site=site)
        elif role:
            nr = nr.filter(vendors=vendors, role=role)
        else:
            nr = nr.filter(vendors='cisco')

    # data = nr.run(task=ios_group_task)
    data = nr.run(task=netmiko_send_commands, commands=commands, conn_timeour=20, global_delay_factor=2)
    result = ResultSerializer(data)
    return result


def result_parsing(result):
    checked_list = list()
    ipa = str()
    for i in result.keys():
        ipa = result[i]['device_ip']
        tmp = dict()
        tmp_str = ''
        for j in result[i].keys():
            '''
            if j == 'device_ip':
                ipa = result[i]['device_ip']
            else:
                ipa = 'unknown'
            '''
            if result[i][j]:
                tmp_str = tmp_str + result[i][j] + '\n'
        cisco_parse = CiscoParse(tmp_str)
        tmp = {
            'ip': ipa,
            'hostname': cisco_parse.hostname(),
            'dev_model': cisco_parse.dev_model(),
            'os_version': cisco_parse.os_ver(),
            'uptime': cisco_parse.uptime(),
            'cpu_idle': cisco_parse.cpu_usage(),
            'mem_free': cisco_parse.mem_usage(),
            'fan': cisco_parse.fan(),
            'temperature': cisco_parse.temperature(),
            'power': cisco_parse.power_supply(),
        }
        checked_list.append(tmp)
    return checked_list


def create_worksheet(excel_file):
    today = datetime.date.today()
    book = load_workbook(excel_file)
    sheet_title = 'report_%s' % (today,)
    header = ['IP', 'Host Name', 'Model', 'Ver.', 'UP Time', 'CPU(% idle)', 'MEM(% idle)', 'Fan', 'Temp.', 'Power']
    try:
        book.create_sheet(title=sheet_title)
        sheet = book[sheet_title]
        # sheet.merge_cells('A1:J1')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        sheet['A1'] = 'Proactive Maintenance Report - %s' % (today,)
        sheet['A1'].font = Font(size=14, bold=True, underline='single')
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.append(header)
        book.save(excel_file)
        book.close()
    except Exception as e:
        print('Some error occurred. \n  %s' % (e,))
        book.close()
    finally:
        book.close()


def report(check_list, excel_file):
    today = datetime.date.today()
    book = load_workbook(excel_file)
    sheet_title = f'report_{today}'
    sheet = book[sheet_title]
    for i in check_list:
        data = [
            i['ip'],
            i['hostname'],
            i['dev_model'],
            i['os_version'],
            i['uptime'],
            i['cpu_idle'],
            i['mem_free'],
            i['fan'],
            i['temperature'],
            i['power'],
        ]
        sheet.append(data)
    book.save(excel_file)
    book.close()
